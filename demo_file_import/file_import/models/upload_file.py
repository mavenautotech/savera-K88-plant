# -*- coding: utf-8 -*-
# Part of Odoo Module Developed by Maven Autotech Pvt. Ltd.

from odoo import models, fields, _
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from odoo.exceptions import UserError
import ftplib
import os
import csv
import shutil
import base64
from selenium import webdriver
import pandas as pd
import glob
import pandas as pd
import os
import glob
import csv
from xlsxwriter.workbook import Workbook


class Uploadfile(models.Model):
    _name = 'file.import'
    _order = 'label_id desc'

    label_id = fields.Char('Label ID')
    print_job_id = fields.Char('Print Job ID')
    ip_address = fields.Char('IP Address')
    verification_type = fields.Char('Verification Type')
    passing_grade_threshold = fields.Char('Passing Grade Threshold')
    date = fields.Char('Date')
    time = fields.Char('Time')
    last_calibration_grade = fields.Char('Last Calibration Date')
    label_numeric_grade = fields.Char('Label Numeric Grade')
    label_grade = fields.Char('Label Grade')
    label_status = fields.Char('Label Status')
    label_failure_reason = fields.Char('Label Failure Reason')
    barcode_id = fields.Char('Barcode ID')
    symbology = fields.Char('Symbology')
    version = fields.Char('Version')
    size = fields.Char('Size')
    numeric_grade = fields.Char('Numeric Grade')
    grade = fields.Char('Grade')
    overall_grade = fields.Char('Overall Grade')
    status = fields.Char('Status')
    scan_line_data_check = fields.Char('Scan Line Data Check')
    xpixel = fields.Char('X Position (Pixel)')
    ypixel = fields.Char('Y Position (Pixel)')
    failure_reason = fields.Char('Failure Reason')
    rmax = fields.Char('R Max')
    rmin = fields.Char('R Min')
    rmin_grade = fields.Char('R Min Grade')
    ec_min = fields.Char('EC Min')
    ec_min_grade = fields.Char('EC Min Grade')
    symbol_contrast = fields.Char('Symbol Contrast')
    symbol_contrast_grade = fields.Char('Symbol Contrast Grade')
    modulation = fields.Char('Modulation')
    modulation_grade = fields.Char('Modulation Grade')
    reflection_margin = fields.Char('Reflectance Margin')
    defects = fields.Char('Defects')
    defects_grade = fields.Char('Defects Grade')
    decodability = fields.Char('Decodability')
    decodability_grade = fields.Char('Decodability Grade')
    decode = fields.Char('Decode')
    quiet_zone = fields.Char('Quiet Zone')
    global_threshold = fields.Char('Global Threshold')
    print_contrast_signal = fields.Char('Print Contrast Signal')
    bargain = fields.Char('Bar Gain')
    axial_nonuniformity = fields.Char('Axial Nonuniformity')
    axial_nonuniformity_grade = fields.Char('Axial Nonuniformity Grade')
    grid_nonuniformity = fields.Char('Grid Nonuniformity')
    grid_nonuniformity_grade = fields.Char('Grid Nonuniformity Grade')
    unused_error = fields.Char('Unused Error Correction')
    unused_error_grade = fields.Char('Unused Error Correction Grade')
    fixed_pattern_damage = fields.Char('Fixed Pattern Damage')
    l1 = fields.Char('L1 (left of L finder)')
    l2 = fields.Char('L2 (bottom of L finder)')
    qzl1 = fields.Char('QZL1 (left quiet zone)')
    qzl2 = fields.Char('QZL2 (bottom quiet zone)')
    octasa = fields.Char('OCTASA (overall clock track and solid area)')
    ag = fields.Char('AG (average grade)')
    segment_a1 = fields.Char('Segment A1')
    segment_a2 = fields.Char('Segment A2')
    segment_a3 = fields.Char('Segment A3')
    segment_b1 = fields.Char('Segment B1')
    segment_b2 = fields.Char('Segment B2')
    segment_c = fields.Char('Segment C')
    format_info = fields.Char('Format Info')
    version_info = fields.Char('Version Info')
    contrast_uniformity = fields.Char('Contrast Uniformity')
    x_growth = fields.Char('X Growth')
    y_growth = fields.Char('Y Growth')
    matrix_size = fields.Char('MatrixSize')
    data = fields.Char('Data')
    encoded_data = fields.Char('Encoded Data')

    def action_done(self):
        path2 = '/home/maven/Music'
        csv_files = glob.glob(path2 + "/*.csv")
        print("^^^^^", csv_files)
        for csvfile in csv_files:
            workbook = Workbook(csvfile[:-4] + '.xlsx')
            worksheet = workbook.add_worksheet()
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.write(r, c, col)
            workbook.close()

        path1 = '/home/maven/Music/'
        list_ = []
        for file_ in path1:
            fileList = glob.glob(path1 + "/*.csv")
            # fileList1 = " ".join(str(x) for x in fileList)
            for x in fileList:
                try:
                    os.remove(x)
                except Exception as e:
                    print("55555", e)


        file1 = list(
            filter(lambda ele: ele if '.xlsx' in ele else None, os.listdir('/home/maven/Music')))
        # print("rrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr")
        for i in file1:
            print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk", i)

            xlsx_file = Path('/home/maven/Music', i)
            print("UU", xlsx_file)
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active
            data = {}

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                # print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA", len(row))
                if len(row) == 70:
                    if i == 0:
                        # print("WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW", len(row), i)
                        data[row[0]] = []
                        data[row[1]] = []
                        data[row[2]] = []
                        data[row[3]] = []
                        data[row[4]] = []
                        data[row[5]] = []
                        data[row[6]] = []
                        data[row[7]] = []
                        data[row[8]] = []
                        data[row[9]] = []
                        data[row[10]] = []
                        data[row[11]] = []
                        data[row[12]] = []
                        data[row[13]] = []
                        data[row[14]] = []
                        data[row[15]] = []
                        data[row[16]] = []
                        data[row[17]] = []
                        data[row[18]] = []
                        data[row[19]] = []
                        data[row[20]] = []
                        data[row[21]] = []
                        data[row[22]] = []
                        data[row[23]] = []
                        data[row[24]] = []
                        data[row[25]] = []
                        data[row[26]] = []
                        data[row[27]] = []
                        data[row[28]] = []
                        data[row[29]] = []
                        data[row[30]] = []
                        data[row[31]] = []
                        data[row[32]] = []
                        data[row[33]] = []
                        data[row[34]] = []
                        data[row[35]] = []
                        data[row[36]] = []
                        data[row[37]] = []
                        data[row[38]] = []
                        data[row[39]] = []
                        data[row[40]] = []
                        data[row[41]] = []
                        data[row[42]] = []
                        data[row[43]] = []
                        data[row[44]] = []
                        data[row[45]] = []
                        data[row[46]] = []
                        data[row[47]] = []
                        data[row[48]] = []
                        data[row[49]] = []
                        data[row[50]] = []
                        data[row[51]] = []
                        data[row[52]] = []
                        data[row[53]] = []
                        data[row[54]] = []
                        data[row[55]] = []
                        data[row[56]] = []
                        data[row[57]] = []
                        data[row[58]] = []
                        data[row[59]] = []
                        data[row[60]] = []
                        data[row[61]] = []
                        data[row[62]] = []
                        data[row[63]] = []
                        data[row[64]] = []
                        data[row[65]] = []
                        data[row[66]] = []
                        data[row[67]] = []
                        data[row[68]] = []
                        data[row[69]] = []

                    else:

                        data['Label ID'].append(row[0])
                        data['Print Job ID'].append(row[1])
                        data['IP Address'].append(row[2])
                        data['Verification Type'].append(row[3])
                        data['Passing Grade Threshold'].append(row[4])
                        data['Date'].append(row[5])
                        data['Time'].append(row[6])
                        data['Last Calibration Date'].append(row[7])
                        data['Label Numeric Grade'].append(row[8])
                        data['Label Grade'].append(row[9])
                        data['Label Status'].append(row[10])
                        data['Label Failure Reason'].append(row[11])
                        data['Barcode ID'].append(row[12])
                        data['Symbology'].append(row[13])
                        data['Version'].append(row[14])
                        data['Size'].append(row[15])
                        data['Numeric Grade'].append(row[16])
                        data['Grade'].append(row[17])
                        data['Overall Grade'].append(row[18])
                        data['Status'].append(row[19])
                        data['Scan Line Data Check'].append(row[20])
                        data['X Position (Pixel)'].append(row[21])
                        data['Y Position (Pixel)'].append(row[22])
                        data['Failure Reason'].append(row[23])
                        data['R Max'].append(row[24])
                        data['R Min'].append(row[25])
                        data['R Min Grade'].append(row[26])
                        data['EC Min'].append(row[27])
                        data['EC Min Grade'].append(row[28])
                        data['Symbol Contrast'].append(row[29])
                        data['Symbol Contrast Grade'].append(row[30])
                        data['Modulation'].append(row[31])
                        data['Modulation Grade'].append(row[32])
                        data['Reflectance Margin'].append(row[33])
                        data['Defects'].append(row[34])
                        data['Defects Grade'].append(row[35])
                        data['Decodability'].append(row[36])
                        data['Decodability Grade'].append(row[37])
                        data['Decode'].append(row[38])
                        data['Quiet Zone'].append(row[39])
                        data['Global Threshold'].append(row[40])
                        data['Print Contrast Signal'].append(row[41])
                        data['Bar Gain'].append(row[42])
                        data['Axial Nonuniformity'].append(row[43])
                        data['Axial Nonuniformity Grade'].append(row[44])
                        data['Grid Nonuniformity'].append(row[45])
                        data['Grid Nonuniformity Grade'].append(row[46])
                        data['Unused Error Correction'].append(row[47])
                        data['Unused Error Correction Grade'].append(row[48])
                        data['Fixed Pattern Damage'].append(row[49])
                        data['L1 (left of L finder)'].append(row[50])
                        data['L2 (bottom of L finder)'].append(row[51])
                        data['QZL1 (left quiet zone)'].append(row[52])
                        data['QZL2 (bottom quiet zone)'].append(row[53])
                        data['OCTASA (overall clock track and solid area)'].append(row[54])
                        data['AG (average grade)'].append(row[55])
                        data['Segment A1'].append(row[56])
                        data['Segment A2'].append(row[57])
                        data['Segment A3'].append(row[58])
                        data['Segment B1'].append(row[59])
                        data['Segment B2'].append(row[60])
                        data['Segment C'].append(row[61])
                        data['Format Info'].append(row[62])
                        data['Version Info'].append(row[63])
                        data['Contrast Uniformity'].append(row[64])
                        data['X Growth'].append(row[65])
                        data['Y Growth'].append(row[66])
                        data['MatrixSize'].append(row[67])
                        data['Data'].append(row[68])
                        data['Encoded Data'].append(row[69])

                        # print("111111111111111111111111111111", data)
                        create_grp = self.env['file.import'].sudo().create(
                            {
                                'label_id': data.get('Label ID')[-1],
                                'print_job_id': data.get('Print Job ID')[-1],
                                'ip_address': data.get('IP Address')[-1],
                                'verification_type': data.get('Verification Type')[-1],
                                'passing_grade_threshold': data.get('Passing Grade Threshold')[-1],
                                'date': data.get('Date')[-1],
                                'time': data.get('Time')[-1],
                                'last_calibration_grade': data.get('Last Calibration Date')[-1],
                                'label_numeric_grade': data.get('Label Numeric Grade')[-1],
                                'label_grade': data.get('Label Grade')[-1],
                                'label_status': data.get('Label Status')[-1],
                                'label_failure_reason': data.get('Label Failure Reason')[-1],
                                'barcode_id': data.get('Barcode ID')[-1],
                                'symbology': data.get('Symbology')[-1],
                                'version': data.get('Version')[-1],
                                'size': data.get('Size')[-1],
                                'numeric_grade': data.get('Numeric Grade')[-1],
                                'grade': data.get('Grade')[-1],
                                'overall_grade': data.get('Overall Grade')[-1],
                                'status': data.get('Status')[-1],
                                'scan_line_data_check': data.get('Scan Line Data Check')[-1],
                                'xpixel': data.get('X Position (Pixel)')[-1],
                                'ypixel': data.get('Y Position (Pixel)')[-1],
                                'failure_reason': data.get('Failure Reason')[-1],
                                'rmax': data.get('R Max')[-1],
                                'rmin': data.get('R Min')[-1],
                                'rmin_grade': data.get('R Min Grade')[-1],
                                'ec_min': data.get('EC Min')[-1],
                                'ec_min_grade': data.get('EC Min Grade')[-1],
                                'symbol_contrast': data.get('Symbol Contrast')[-1],
                                'symbol_contrast_grade': data.get('Symbol Contrast Grade')[-1],
                                'modulation': data.get('Modulation')[-1],
                                'modulation_grade': data.get('Modulation Grade')[-1],
                                'reflection_margin': data.get('Reflectance Margin')[-1],
                                'defects': data.get('Defects')[-1],
                                'defects_grade': data.get('Defects Grade')[-1],
                                'decodability': data.get('Decodability')[-1],
                                'decodability_grade': data.get('Decodability Grade')[-1],
                                'decode': data.get('Decode')[-1],
                                'quiet_zone': data.get('Quiet Zone')[-1],
                                'global_threshold': data.get('Global Threshold')[-1],
                                'print_contrast_signal': data.get('Print Contrast Signal')[-1],
                                'bargain': data.get('Bar Gain')[-1],
                                'axial_nonuniformity': data.get('Axial Nonuniformity')[-1],
                                'axial_nonuniformity_grade': data.get('Axial Nonuniformity Grade')[-1],
                                'grid_nonuniformity': data.get('Grid Nonuniformity')[-1],
                                'grid_nonuniformity_grade': data.get('Grid Nonuniformity Grade')[-1],
                                'unused_error': data.get('Unused Error Correction')[-1],
                                'unused_error_grade': data.get('Unused Error Correction Grade')[-1],
                                'fixed_pattern_damage': data.get('Fixed Pattern Damage')[-1],
                                'l1': data.get('L1 (left of L finder)')[-1],
                                'l2': data.get('L2 (bottom of L finder)')[-1],
                                'qzl1': data.get('QZL1 (left quiet zone)')[-1],
                                'qzl2': data.get('QZL2 (bottom quiet zone)')[-1],
                                'octasa': data.get('OCTASA (overall clock track and solid area)')[-1],
                                'ag': data.get('AG (average grade)')[-1],
                                'segment_a1': data.get('Segment A1')[-1],
                                'segment_a2': data.get('Segment A2')[-1],
                                'segment_a3': data.get('Segment A3')[-1],
                                'segment_b1': data.get('Segment B1')[-1],
                                'segment_b2': data.get('Segment B2')[-1],
                                'segment_c': data.get('Segment C')[-1],
                                'format_info': data.get('Format Info')[-1],
                                'version_info': data.get('Version Info')[-1],
                                'contrast_uniformity': data.get('Contrast Uniformity')[-1],
                                'x_growth': data.get('X Growth')[-1],
                                'y_growth': data.get('Y Growth')[-1],
                                'matrix_size': data.get('MatrixSize')[-1],
                                'data': data.get('Data')[-1],
                                'encoded_data': data.get('Encoded Data')[-1],

                            })
                        # print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA", data)

                else:
                    if i == 0:
                        # print("QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQ33333333", len(row))
                        data[row[0]] = []
                        data[row[1]] = []
                        data[row[2]] = []
                        data[row[3]] = []
                        data[row[4]] = []
                        data[row[5]] = []
                        data[row[6]] = []
                        data[row[7]] = []
                        data[row[8]] = []
                        data[row[9]] = []
                        data[row[10]] = []
                        data[row[11]] = []
                        data[row[12]] = []
                        data[row[13]] = []
                        data[row[14]] = []
                        data[row[15]] = []
                        data[row[16]] = []
                        data[row[17]] = []
                        data[row[18]] = []
                        data[row[19]] = []
                        data[row[20]] = []
                        data[row[21]] = []
                        data[row[22]] = []
                        data[row[23]] = []
                        data[row[24]] = []
                        data[row[25]] = []
                        data[row[26]] = []
                        data[row[27]] = []
                        data[row[28]] = []
                        data[row[29]] = []
                        data[row[30]] = []
                        data[row[31]] = []
                        data[row[32]] = []
                        data[row[33]] = []
                        data[row[34]] = []
                        data[row[35]] = []
                        data[row[36]] = []
                        data[row[37]] = []
                        data[row[38]] = []
                        data[row[39]] = []
                        data[row[40]] = []
                        data[row[41]] = []
                        data[row[42]] = []
                        data[row[43]] = []
                        data[row[44]] = []
                        data[row[45]] = []
                        data[row[46]] = []
                        data[row[47]] = []
                        data[row[48]] = []
                        data[row[49]] = []
                        data[row[50]] = []
                        data[row[51]] = []
                        data[row[52]] = []
                        data[row[53]] = []
                        data[row[54]] = []
                        data[row[55]] = []
                        data[row[56]] = []
                        data[row[57]] = []
                        data[row[58]] = []
                        data[row[59]] = []
                        data[row[60]] = []
                        data[row[61]] = []
                        data[row[62]] = []
                        data[row[63]] = []
                        data[row[64]] = []
                        data[row[65]] = []
                        data[row[66]] = []
                        data[row[67]] = []
                        data[row[68]] = []
                        data[row[69]] = []

                    else:
                        # print("SSSSSSSSSSSSSSSSSSSSSSSSSSSSSSS", len(row))
                        data['Label ID'].append(row[0])
                        data['Print Job ID'].append(row[1])
                        data['IP Address'].append(row[2])
                        data['Verification Type'].append(row[3])
                        data['Passing Grade Threshold'].append(row[4])
                        data['Date'].append(row[5])
                        data['Time'].append(row[6])
                        data['Last Calibration Date'].append(row[7])
                        data['Label Numeric Grade'].append(row[8])
                        data['Label Grade'].append(row[9])
                        data['Label Status'].append(row[10])
                        data['Label Failure Reason'].append(row[11])
                        data['Barcode ID'].append(row[12])
                        data['Symbology'].append(row[13])
                        data['Version'].append(row[14])
                        data['Size'].append(row[15])
                        data['Numeric Grade'].append(row[16])
                        data['Grade'].append(row[17])
                        data['Overall Grade'].append(row[18])
                        data['Status'].append(row[19])
                        data['Scan Line Data Check'].append(row[20])
                        data['X Position (Pixel)'].append(row[21])
                        data['Y Position (Pixel)'].append(row[22])
                        data['Failure Reason'].append(row[23])
                        data['R Max'].append(row[24])
                        data['R Min'].append(row[25])
                        data['R Min Grade'].append(row[26])
                        data['EC Min'].append(row[27])
                        data['EC Min Grade'].append(row[28])
                        data['Symbol Contrast'].append(row[29])
                        data['Symbol Contrast Grade'].append(row[30])
                        data['Modulation'].append(row[31])
                        data['Modulation Grade'].append(row[32])
                        data['Reflectance Margin'].append(row[33])
                        data['Defects'].append(row[34])
                        data['Defects Grade'].append(row[35])
                        data['Decodability'].append(row[36])
                        data['Decodability Grade'].append(row[37])
                        data['Decode'].append(row[38])
                        data['Quiet Zone'].append(row[39])
                        data['Global Threshold'].append(row[40])
                        data['Print Contrast Signal'].append(row[41])
                        data['Bar Gain'].append(row[42])
                        data['Axial Nonuniformity'].append(row[43])
                        data['Axial Nonuniformity Grade'].append(row[44])
                        data['Grid Nonuniformity'].append(row[45])
                        data['Grid Nonuniformity Grade'].append(row[46])
                        data['Unused Error Correction'].append(row[47])
                        data['Unused Error Correction Grade'].append(row[48])
                        data['Fixed Pattern Damage'].append(row[49])
                        data['L1 (left of L finder)'].append(row[50])
                        data['L2 (bottom of L finder)'].append(row[51])
                        data['QZL1 (left quiet zone)'].append(row[52])
                        data['QZL2 (bottom quiet zone)'].append(row[53])
                        data['OCTASA (overall clock track and solid area)'].append(row[54])
                        data['AG (average grade)'].append(row[55])
                        data['Segment A1'].append(row[56])
                        data['Segment A2'].append(row[57])
                        data['Segment A3'].append(row[58])
                        data['Segment B1'].append(row[59])
                        data['Segment B2'].append(row[60])
                        data['Segment C'].append(row[61])
                        data['Format Info'].append(row[62])
                        data['Version Info'].append(row[63])
                        data['Contrast Uniformity'].append(row[64])
                        data['X Growth'].append(row[65])
                        data['Y Growth'].append(row[66])
                        data['MatrixSize'].append(row[67])
                        data['Data'].append(row[68])
                        data['Encoded Data'].append(row[69])

                        create_grp = self.env['file.import'].sudo().create(
                            {
                                'label_id': data.get('Label ID')[-1],
                                'print_job_id': data.get('Print Job ID')[-1],
                                'ip_address': data.get('IP Address')[-1],
                                'verification_type': data.get('Verification Type')[-1],
                                'passing_grade_threshold': data.get('Passing Grade Threshold')[-1],
                                'date': data.get('Date')[-1],
                                'time': data.get('Time')[-1],
                                'last_calibration_grade': data.get('Last Calibration Date')[-1],
                                'label_numeric_grade': data.get('Label Numeric Grade')[-1],
                                'label_grade': data.get('Label Grade')[-1],
                                'label_status': data.get('Label Status')[-1],
                                'label_failure_reason': data.get('Label Failure Reason')[-1],
                                'barcode_id': data.get('Barcode ID')[-1],
                                'symbology': data.get('Symbology')[-1],
                                'version': data.get('Version')[-1],
                                'size': data.get('Size')[-1],
                                'numeric_grade': data.get('Numeric Grade')[-1],
                                'grade': data.get('Grade')[-1],
                                'overall_grade': data.get('Overall Grade')[-1],
                                'status': data.get('Status')[-1],
                                'scan_line_data_check': data.get('Scan Line Data Check')[-1],
                                'xpixel': data.get('X Position (Pixel)')[-1],
                                'ypixel': data.get('Y Position (Pixel)')[-1],
                                'failure_reason': data.get('Failure Reason')[-1],
                                'rmax': data.get('R Max')[-1],
                                'rmin': data.get('R Min')[-1],
                                'rmin_grade': data.get('R Min Grade')[-1],
                                'ec_min': data.get('EC Min')[-1],
                                'ec_min_grade': data.get('EC Min Grade')[-1],
                                'symbol_contrast': data.get('Symbol Contrast')[-1],
                                'symbol_contrast_grade': data.get('Symbol Contrast Grade')[-1],
                                'modulation': data.get('Modulation')[-1],
                                'modulation_grade': data.get('Modulation Grade')[-1],
                                'reflection_margin': data.get('Reflectance Margin')[-1],
                                'defects': data.get('Defects')[-1],
                                'defects_grade': data.get('Defects Grade')[-1],
                                'decodability': data.get('Decodability')[-1],
                                'decodability_grade': data.get('Decodability Grade')[-1],
                                'decode': data.get('Decode')[-1],
                                'quiet_zone': data.get('Quiet Zone')[-1],
                                'global_threshold': data.get('Global Threshold')[-1],
                                'print_contrast_signal': data.get('Print Contrast Signal')[-1],
                                'bargain': data.get('Bar Gain')[-1],
                                'axial_nonuniformity': data.get('Axial Nonuniformity')[-1],
                                'axial_nonuniformity_grade': data.get('Axial Nonuniformity Grade')[-1],
                                'grid_nonuniformity': data.get('Grid Nonuniformity')[-1],
                                'grid_nonuniformity_grade': data.get('Grid Nonuniformity Grade')[-1],
                                'unused_error': data.get('Unused Error Correction')[-1],
                                'unused_error_grade': data.get('Unused Error Correction Grade')[-1],
                                'fixed_pattern_damage': data.get('Fixed Pattern Damage')[-1],
                                'l1': data.get('L1 (left of L finder)')[-1],
                                'l2': data.get('L2 (bottom of L finder)')[-1],
                                'qzl1': data.get('QZL1 (left quiet zone)')[-1],
                                'qzl2': data.get('QZL2 (bottom quiet zone)')[-1],
                                'octasa': data.get('OCTASA (overall clock track and solid area)')[-1],
                                'ag': data.get('AG (average grade)')[-1],
                                'segment_a1': data.get('Segment A1')[-1],
                                'segment_a2': data.get('Segment A2')[-1],
                                'segment_a3': data.get('Segment A3')[-1],
                                'segment_b1': data.get('Segment B1')[-1],
                                'segment_b2': data.get('Segment B2')[-1],
                                'segment_c': data.get('Segment C')[-1],
                                'format_info': data.get('Format Info')[-1],
                                'version_info': data.get('Version Info')[-1],
                                'contrast_uniformity': data.get('Contrast Uniformity')[-1],
                                'x_growth': data.get('X Growth')[-1],
                                'y_growth': data.get('Y Growth')[-1],
                                'matrix_size': data.get('MatrixSize')[-1],
                                'data': data.get('Data')[-1],
                                'encoded_data': data.get('Encoded Data')[-1],

                            })
                        path2 = '/home/maven/Music/'
                        list_ = []
                        for file_ in path2:
                            fileList1 = glob.glob(path2 + "/*.xlsx")
                            for x in fileList1:
                                try:
                                    os.remove(x)
                                except Exception as e:
                                    print(e)
                        # print("BBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBB", data)

#     def action_done(self):
#         # ftp_conf_obj = self.env['ftp.config'].sudo().search([])
#
#         # try:
#         # ftp_user_id = ftp_conf_obj[-1].ftp_user_id
#         # ftp_password = ftp_conf_obj[-1].ftp_password
#         # ftp_path_testing = ftp_conf_obj[-1].ftp_path_testing
#         # print("RRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR", ftp_path_testing)
#         # ip = "192.168.1.4"
#         # id = ftp_user_id.strip()
#         # ps = ftp_password.strip()
#         # ftp = ftplib.FTP(ip)
#         # ftp.login(id, ps)
#         # print("++++++++++++++++++++++++++++++++++++++++++++++++++++++", ftp.login)
#         # except:
#         #     raise UserError("Could not login to FTP! Please check FTP credentials")
#         # try:
#         # path = "\Extra files"
#         # ftp.cwd(path)
#         # file = ftp.nlst()[0]
#         # path = ftp_path_testing
#         # print("PPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPp", path)
#         # ftp.cwd(path)
#         # file = ftp.nlst()[0]
#
#         #     fwp = os.path.join("//home/maven/Music/", file)
#         #     with open(fwp, 'wb') as f:
#         #         ftp.retrbinary('RETR ' + file, f.write)
#         #     ftp.close()
#         # except:
#         #     raise UserError(
#         #         "Some thing went wrong while fetching file from FTP. Please check file location/folder name")
#
#         path2 = '/home/maven/Music'
#         csv_files = glob.glob(path2 + "/*.csv")
#         print("^^^^^", csv_files)
#         for file2 in csv_files:
#             df_list = pd.read_csv(file2)
#             GFG = pd.ExcelWriter(file2 + '.xlsx')
#             # print(":::::::", df_list)
#             df_list.to_excel(GFG, index=False, header=True)
#             # print("________", df_list)
#             GFG.close()
#
#         path1 = '/home/maven/Music/'
#         list_ = []
#         for file_ in path1:
#             fileList = glob.glob(path1 + "/*.csv")
#             fileList1 = " ".join(str(x) for x in fileList)
#             try:
#                 os.remove(fileList1)
#             except Exception as e:
#                 print("55555", e)
#
#         file1 = list(
#             filter(lambda ele: ele if '.xlsx' in ele else None, os.listdir('/home/maven/Music')))
#         xlsx_file = Path('/home/maven/Music', file1[0])
#         print("UU", xlsx_file)
#         wb_obj = openpyxl.load_workbook(xlsx_file)
#         sheet = wb_obj.active
#         data = {}
#
#         for i, row in enumerate(sheet.iter_rows(values_only=True)):
#             print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA", len(row))
#             if len(row) == 70:
#                 if i == 0:
#                     print("WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW", len(row), i)
#                     data[row[0]] = []
#                     data[row[1]] = []
#                     data[row[2]] = []
#                     data[row[3]] = []
#                     data[row[4]] = []
#                     data[row[5]] = []
#                     # data[row[6]] = []
#                     # data[row[7]] = []
#                     # data[row[8]] = []
#                     # data[row[9]] = []
#                     # data[row[10]] = []
#                     # data[row[11]] = []
#                     # data[row[12]] = []
#                     # data[row[13]] = []
#                     # data[row[14]] = []
#                     # data[row[15]] = []
#                     # data[row[16]] = []
#                     # data[row[17]] = []
#                     # data[row[18]] = []
#                     # data[row[19]] = []
#                     # data[row[20]] = []
#                     # data[row[21]] = []
#                     # data[row[22]] = []
#                     # data[row[23]] = []
#                     # data[row[24]] = []
#                     # data[row[25]] = []
#                     # data[row[26]] = []
#                     # data[row[27]] = []
#                     # data[row[28]] = []
#                     # data[row[29]] = []
#                     # data[row[30]] = []
#                     # data[row[31]] = []
#                     # data[row[32]] = []
#                     # data[row[33]] = []
#                     # data[row[34]] = []
#                     # data[row[35]] = []
#                     # data[row[36]] = []
#                     # data[row[37]] = []
#                     # data[row[38]] = []
#                     # data[row[39]] = []
#                     # data[row[40]] = []
#                     # data[row[41]] = []
#                     # data[row[42]] = []
#                     # data[row[43]] = []
#                     # data[row[44]] = []
#                     # data[row[45]] = []
#                     # data[row[46]] = []
#                     # data[row[47]] = []
#                     # data[row[48]] = []
#                     # data[row[49]] = []
#                     # data[row[50]] = []
#                     # data[row[51]] = []
#                     # data[row[52]] = []
#                     # data[row[53]] = []
#                     # data[row[54]] = []
#                     # data[row[55]] = []
#                     # data[row[56]] = []
#                     # data[row[57]] = []
#                     # data[row[58]] = []
#                     # data[row[59]] = []
#                     # data[row[60]] = []
#                     # data[row[61]] = []
#                     # data[row[62]] = []
#                     # data[row[63]] = []
#                     # data[row[64]] = []
#                     # data[row[65]] = []
#                     # data[row[66]] = []
#                     # data[row[67]] = []
#                     # data[row[68]] = []
#                     # data[row[69]] = []
#
#                 else:
#                     try:
#                         data['Label ID'].append(row[0])
#                         data['Print Job ID'].append(row[1])
#                         data['IP Address'].append(row[2])
#                         data['Verification Type'].append(row[3])
#                         data['Passing Grade Threshold'].append(row[4])
#                         data['Date'].append(row[5])
#                         # data['Time'].append(row[6])
#                         # data['Last Calibration Date'].append(row[7])
#                         # data['Label Numeric Grade'].append(row[8])
#                         # data['Label Grade'].append(row[9])
#                         # data['Label Status'].append(row[10])
#                         # data['Label Failure Reason'].append(row[11])
#                         # data['Barcode ID'].append(row[12])
#                         # data['Symbology'].append(row[13])
#                         # data['Version'].append(row[14])
#                         # data['Size'].append(row[15])
#                         # data['Numeric Grade'].append(row[16])
#                         # data['Grade'].append(row[17])
#                         # data['Overall Grade'].append(row[18])
#                         # data['Status'].append(row[19])
#                         # data['Scan Line Data Check'].append(row[20])
#                         # data['X Position (Pixel)'].append(row[21])
#                         # data['Y Position (Pixel)'].append(row[22])
#                         # data['Failure Reason'].append(row[23])
#                         # data['R Max'].append(row[24])
#                         # data['R Min'].append(row[25])
#                         # data['R Min Grade'].append(row[26])
#                         # data['EC Min'].append(row[27])
#                         # data['EC Min Grade'].append(row[28])
#                         # data['Symbol Contrast'].append(row[29])
#                         # data['Symbol Contrast Grade'].append(row[30])
#                         # data['Modulation'].append(row[31])
#                         # data['Modulation Grade'].append(row[32])
#                         # data['Reflectance Margin'].append(row[33])
#                         # data['Defects'].append(row[34])
#                         # data['Defects Grade'].append(row[35])
#                         # data['Decodability'].append(row[36])
#                         # data['Decodability Grade'].append(row[37])
#                         # data['Decode'].append(row[38])
#                         # data['Quiet Zone'].append(row[39])
#                         # data['Global Threshold'].append(row[40])
#                         # data['Print Contrast Signal'].append(row[41])
#                         # data['Bar Gain'].append(row[42])
#                         # data['Axial Nonuniformity'].append(row[43])
#                         # data['Axial Nonuniformity Grade'].append(row[44])
#                         # data['Grid Nonuniformity'].append(row[45])
#                         # data['Grid Nonuniformity Grade'].append(row[46])
#                         # data['Unused Error Correction'].append(row[47])
#                         # data['Unused Error Correction Grade'].append(row[48])
#                         # data['Fixed Pattern Damage'].append(row[49])
#                         # data['L1 (left of L finder)'].append(row[50])
#                         # data['L2 (bottom of L finder)'].append(row[51])
#                         # data['QZL1 (left quiet zone)'].append(row[52])
#                         # data['QZL2 (bottom quiet zone)'].append(row[53])
#                         # data['OCTASA (overall clock track and solid area)'].append(row[54])
#                         # data['AG (average grade)'].append(row[55])
#                         # data['Segment A1'].append(row[56])
#                         # data['Segment A2'].append(row[57])
#                         # data['Segment A3'].append(row[58])
#                         # data['Segment B1'].append(row[59])
#                         # data['Segment B2'].append(row[60])
#                         # data['Segment C'].append(row[61])
#                         # data['Format Info'].append(row[62])
#                         # data['Version Info'].append(row[63])
#                         # data['Contrast Uniformity'].append(row[64])
#                         # data['X Growth'].append(row[65])
#                         # data['Y Growth'].append(row[66])
#                         # data['MatrixSize'].append(row[67])
#                         # data['Data'].append(row[68])
#                         # data['Encoded Data'].append(row[69])
#                         print("111111111111111111111111111111", data)
#                         create_grp = self.env['file.import'].sudo().create(
#                             {
#                                 'label_id': data.get('Label ID')[-1],
#                                 'print_job_id': data.get('Print Job ID')[-1],
#                                 'ip_address': data.get('IP Address')[-1],
#                                 'verification_type': data.get('Verification Type')[-1],
#                                 'passing_grade_threshold': data.get('Passing Grade Threshold')[-1],
#                                 'date': data.get('Date')[-1],
#                                 # 'time': data.get('Time')[-1],
#                                 # 'last_calibration_grade': data.get('Last Calibration Date')[-1],
#                                 # 'label_numeric_grade': data.get('Label Numeric Grade')[-1],
#                                 # 'label_grade': data.get('Label Grade')[-1],
#                                 # 'label_status': data.get('Label Status')[-1],
#                                 # 'label_failure_reason': data.get('Label Failure Reason')[-1],
#                                 # 'barcode_id': data.get('Barcode ID')[-1],
#                                 # 'symbology': data.get('Symbology')[-1],
#                                 # 'version': data.get('Version')[-1],
#                                 # 'size': data.get('Size')[-1],
#                                 # 'numeric_grade': data.get('Numeric Grade')[-1],
#                                 # 'grade': data.get('Grade')[-1],
#                                 # 'overall_grade': data.get('Overall Grade')[-1],
#                                 # 'status': data.get('Status')[-1],
#                                 # 'scan_line_data_check': data.get('Scan Line Data Check')[-1],
#                                 # 'xpixel': data.get('X Position (Pixel)')[-1],
#                                 # 'ypixel': data.get('Y Position (Pixel)')[-1],
#                                 # 'failure_reason': data.get('Failure Reason')[-1],
#                                 # 'rmax': data.get('R Max')[-1],
#                                 # 'rmin': data.get('R Min')[-1],
#                                 # 'rmin_grade': data.get('R Min Grade')[-1],
#                                 # 'ec_min': data.get('EC Min')[-1],
#                                 # 'ec_min_grade': data.get('EC Min Grade')[-1],
#                                 # 'symbol_contrast': data.get('Symbol Contrast')[-1],
#                                 # 'symbol_contrast_grade': data.get('Symbol Contrast Grade')[-1],
#                                 # 'modulation': data.get('Modulation')[-1],
#                                 # 'modulation_grade': data.get('Modulation Grade')[-1],
#                                 # 'reflection_margin': data.get('Reflectance Margin')[-1],
#                                 # 'defects': data.get('Defects')[-1],
#                                 # 'defects_grade': data.get('Defects Grade')[-1],
#                                 # 'decodability': data.get('Decodability')[-1],
#                                 # 'decodability_grade': data.get('Decodability Grade')[-1],
#                                 # 'decode': data.get('Decode')[-1],
#                                 # 'quiet_zone': data.get('Quiet Zone')[-1],
#                                 # 'global_threshold': data.get('Global Threshold')[-1],
#                                 # 'print_contrast_signal': data.get('Print Contrast Signal')[-1],
#                                 # 'bargain': data.get('Bar Gain')[-1],
#                                 # 'axial_nonuniformity': data.get('Axial Nonuniformity')[-1],
#                                 # 'axial_nonuniformity_grade': data.get('Axial Nonuniformity Grade')[-1],
#                                 # 'grid_nonuniformity': data.get('Grid Nonuniformity')[-1],
#                                 # 'grid_nonuniformity_grade': data.get('Grid Nonuniformity Grade')[-1],
#                                 # 'unused_error': data.get('Unused Error Correction')[-1],
#                                 # 'unused_error_grade': data.get('Unused Error Correction Grade')[-1],
#                                 # 'fixed_pattern_damage': data.get('Fixed Pattern Damage')[-1],
#                                 # 'l1': data.get('L1 (left of L finder)')[-1],
#                                 # 'l2': data.get('L2 (bottom of L finder)')[-1],
#                                 # 'qzl1': data.get('QZL1 (left quiet zone)')[-1],
#                                 # 'qzl2': data.get('QZL2 (bottom quiet zone)')[-1],
#                                 # 'octasa': data.get('OCTASA (overall clock track and solid area)')[-1],
#                                 # 'ag': data.get('AG (average grade)')[-1],
#                                 # 'segment_a1': data.get('Segment A1')[-1],
#                                 # 'segment_a2': data.get('Segment A2')[-1],
#                                 # 'segment_a3': data.get('Segment A3')[-1],
#                                 # 'segment_b1': data.get('Segment B1')[-1],
#                                 # 'segment_b2': data.get('Segment B2')[-1],
#                                 # 'segment_c': data.get('Segment C')[-1],
#                                 # 'format_info': data.get('Format Info')[-1],
#                                 # 'version_info': data.get('Version Info')[-1],
#                                 # 'contrast_uniformity': data.get('Contrast Uniformity')[-1],
#                                 # 'x_growth': data.get('X Growth')[-1],
#                                 # 'y_growth': data.get('Y Growth')[-1],
#                                 # 'matrix_size': data.get('MatrixSize')[-1],
#                                 # 'data': data.get('Data')[-1],
#                                 # 'encoded_data': data.get('Encoded Data')[-1],
#
#                             })
#                         print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA", data)
#                     except:
#                         raise UserError(_("Column header mismatched !"))
#             else:
#                 if i == 0:
#                     print("QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQ33333333", len(row))
#                     data[row[0]] = []
#                     data[row[1]] = []
#                     data[row[2]] = []
#                     data[row[3]] = []
#                     data[row[4]] = []
#                     data[row[5]] = []
#                     # data[row[6]] = []
#                     # data[row[7]] = []
#                     # data[row[8]] = []
#                     # data[row[9]] = []
#                     # data[row[10]] = []
#                     # data[row[11]] = []
#                     # data[row[12]] = []
#                     # data[row[13]] = []
#                     # data[row[14]] = []
#                     # data[row[15]] = []
#                     # data[row[16]] = []
#                     # data[row[17]] = []
#                     # data[row[18]] = []
#                     # data[row[19]] = []
#                     # data[row[20]] = []
#                     # data[row[21]] = []
#                     # data[row[22]] = []
#                     # data[row[23]] = []
#                     # data[row[24]] = []
#                     # data[row[25]] = []
#                     # data[row[26]] = []
#                     # data[row[27]] = []
#                     # data[row[28]] = []
#                     # data[row[29]] = []
#                     # data[row[30]] = []
#                     # data[row[31]] = []
#                     # data[row[32]] = []
#                     # data[row[33]] = []
#                     # data[row[34]] = []
#                     # data[row[35]] = []
#                     # data[row[36]] = []
#                     # data[row[37]] = []
#                     # data[row[38]] = []
#                     # data[row[39]] = []
#                     # data[row[40]] = []
#                     # data[row[41]] = []
#                     # data[row[42]] = []
#                     # data[row[43]] = []
#                     # data[row[44]] = []
#                     # data[row[45]] = []
#                     # data[row[46]] = []
#                     # data[row[47]] = []
#                     # data[row[48]] = []
#                     # data[row[49]] = []
#                     # data[row[50]] = []
#                     # data[row[51]] = []
#                     # data[row[52]] = []
#                     # data[row[53]] = []
#                     # data[row[54]] = []
#                     # data[row[55]] = []
#                     # data[row[56]] = []
#                     # data[row[57]] = []
#                     # data[row[58]] = []
#                     # data[row[59]] = []
#                     # data[row[60]] = []
#                     # data[row[61]] = []
#                     # data[row[62]] = []
#                     # data[row[63]] = []
#                     # data[row[64]] = []
#                     # data[row[65]] = []
#                     # data[row[66]] = []
#                     # data[row[67]] = []
#                     # data[row[68]] = []
#                     # data[row[69]] = []
#
#                 else:
#                     print("SSSSSSSSSSSSSSSSSSSSSSSSSSSSSSS", len(row))
#                     data['Label ID'].append(row[0])
#                     data['Print Job ID'].append(row[1])
#                     data['IP Address'].append(row[2])
#                     data['Verification Type'].append(row[3])
#                     data['Passing Grade Threshold'].append(row[4])
#                     data['Date'].append(row[5])
#                     # data['Time'].append(row[6])
#                     # data['Last Calibration Date'].append(row[7])
#                     # data['Label Numeric Grade'].append(row[8])
#                     # data['Label Grade'].append(row[9])
#                     # data['Label Status'].append(row[10])
#                     # data['Label Failure Reason'].append(row[11])
#                     # data['Barcode ID'].append(row[12])
#                     # data['Symbology'].append(row[13])
#                     # data['Version'].append(row[14])
#                     # data['Size'].append(row[15])
#                     # data['Numeric Grade'].append(row[16])
#                     # data['Grade'].append(row[17])
#                     # data['Overall Grade'].append(row[18])
#                     # data['Status'].append(row[19])
#                     # data['Scan Line Data Check'].append(row[20])
#                     # data['X Position (Pixel)'].append(row[21])
#                     # data['Y Position (Pixel)'].append(row[22])
#                     # data['Failure Reason'].append(row[23])
#                     # data['R Max'].append(row[24])
#                     # data['R Min'].append(row[25])
#                     # data['R Min Grade'].append(row[26])
#                     # data['EC Min'].append(row[27])
#                     # data['EC Min Grade'].append(row[28])
#                     # data['Symbol Contrast'].append(row[29])
#                     # data['Symbol Contrast Grade'].append(row[30])
#                     # data['Modulation'].append(row[31])
#                     # data['Modulation Grade'].append(row[32])
#                     # data['Reflectance Margin'].append(row[33])
#                     # data['Defects'].append(row[34])
#                     # data['Defects Grade'].append(row[35])
#                     # data['Decodability'].append(row[36])
#                     # data['Decodability Grade'].append(row[37])
#                     # data['Decode'].append(row[38])
#                     # data['Quiet Zone'].append(row[39])
#                     # data['Global Threshold'].append(row[40])
#                     # data['Print Contrast Signal'].append(row[41])
#                     # data['Bar Gain'].append(row[42])
#                     # data['Axial Nonuniformity'].append(row[43])
#                     # data['Axial Nonuniformity Grade'].append(row[44])
#                     # data['Grid Nonuniformity'].append(row[45])
#                     # data['Grid Nonuniformity Grade'].append(row[46])
#                     # data['Unused Error Correction'].append(row[47])
#                     # data['Unused Error Correction Grade'].append(row[48])
#                     # data['Fixed Pattern Damage'].append(row[49])
#                     # data['L1 (left of L finder)'].append(row[50])
#                     # data['L2 (bottom of L finder)'].append(row[51])
#                     # data['QZL1 (left quiet zone)'].append(row[52])
#                     # data['QZL2 (bottom quiet zone)'].append(row[53])
#                     # data['OCTASA (overall clock track and solid area)'].append(row[54])
#                     # data['AG (average grade)'].append(row[55])
#                     # data['Segment A1'].append(row[56])
#                     # data['Segment A2'].append(row[57])
#                     # data['Segment A3'].append(row[58])
#                     # data['Segment B1'].append(row[59])
#                     # data['Segment B2'].append(row[60])
#                     # data['Segment C'].append(row[61])
#                     # data['Format Info'].append(row[62])
#                     # data['Version Info'].append(row[63])
#                     # data['Contrast Uniformity'].append(row[64])
#                     # data['X Growth'].append(row[65])
#                     # data['Y Growth'].append(row[66])
#                     # data['MatrixSize'].append(row[67])
#                     # data['Data'].append(row[68])
#                     # data['Encoded Data'].append(row[69])
#
#                     create_grp = self.env['file.import'].sudo().create(
#                         {
#                             'label_id': data.get('Label ID')[-1],
#                             'print_job_id': data.get('Print Job ID')[-1],
#                             'ip_address': data.get('IP Address')[-1],
#                             'verification_type': data.get('Verification Type')[-1],
#                             'passing_grade_threshold': data.get('Passing Grade Threshold')[-1],
#                             'date': data.get('Date')[-1],
#                             # 'time': data.get('Time')[-1],
#                             # 'last_calibration_grade': data.get('Last Calibration Date')[-1],
#                             # 'label_numeric_grade': data.get('Label Numeric Grade')[-1],
#                             # 'label_grade': data.get('Label Grade')[-1],
#                             # 'label_status': data.get('Label Status')[-1],
#                             # 'label_failure_reason': data.get('Label Failure Reason')[-1],
#                             # 'barcode_id': data.get('Barcode ID')[-1],
#                             # 'symbology': data.get('Symbology')[-1],
#                             # 'version': data.get('Version')[-1],
#                             # 'size': data.get('Size')[-1],
#                             # 'numeric_grade': data.get('Numeric Grade')[-1],
#                             # 'grade': data.get('Grade')[-1],
#                             # 'overall_grade': data.get('Overall Grade')[-1],
#                             # 'status': data.get('Status')[-1],
#                             # 'scan_line_data_check': data.get('Scan Line Data Check')[-1],
#                             # 'xpixel': data.get('X Position (Pixel)')[-1],
#                             # 'ypixel': data.get('Y Position (Pixel)')[-1],
#                             # 'failure_reason': data.get('Failure Reason')[-1],
#                             # 'rmax': data.get('R Max')[-1],
#                             # 'rmin': data.get('R Min')[-1],
#                             # 'rmin_grade': data.get('R Min Grade')[-1],
#                             # 'ec_min': data.get('EC Min')[-1],
#                             # 'ec_min_grade': data.get('EC Min Grade')[-1],
#                             # 'symbol_contrast': data.get('Symbol Contrast')[-1],
#                             # 'symbol_contrast_grade': data.get('Symbol Contrast Grade')[-1],
#                             # 'modulation': data.get('Modulation')[-1],
#                             # 'modulation_grade': data.get('Modulation Grade')[-1],
#                             # 'reflection_margin': data.get('Reflectance Margin')[-1],
#                             # 'defects': data.get('Defects')[-1],
#                             # 'defects_grade': data.get('Defects Grade')[-1],
#                             # 'decodability': data.get('Decodability')[-1],
#                             # 'decodability_grade': data.get('Decodability Grade')[-1],
#                             # 'decode': data.get('Decode')[-1],
#                             # 'quiet_zone': data.get('Quiet Zone')[-1],
#                             # 'global_threshold': data.get('Global Threshold')[-1],
#                             # 'print_contrast_signal': data.get('Print Contrast Signal')[-1],
#                             # 'bargain': data.get('Bar Gain')[-1],
#                             # 'axial_nonuniformity': data.get('Axial Nonuniformity')[-1],
#                             # 'axial_nonuniformity_grade': data.get('Axial Nonuniformity Grade')[-1],
#                             # 'grid_nonuniformity': data.get('Grid Nonuniformity')[-1],
#                             # 'grid_nonuniformity_grade': data.get('Grid Nonuniformity Grade')[-1],
#                             # 'unused_error': data.get('Unused Error Correction')[-1],
#                             # 'unused_error_grade': data.get('Unused Error Correction Grade')[-1],
#                             # 'fixed_pattern_damage': data.get('Fixed Pattern Damage')[-1],
#                             # 'l1': data.get('L1 (left of L finder)')[-1],
#                             # 'l2': data.get('L2 (bottom of L finder)')[-1],
#                             # 'qzl1': data.get('QZL1 (left quiet zone)')[-1],
#                             # 'qzl2': data.get('QZL2 (bottom quiet zone)')[-1],
#                             # 'octasa': data.get('OCTASA (overall clock track and solid area)')[-1],
#                             # 'ag': data.get('AG (average grade)')[-1],
#                             # 'segment_a1': data.get('Segment A1')[-1],
#                             # 'segment_a2': data.get('Segment A2')[-1],
#                             # 'segment_a3': data.get('Segment A3')[-1],
#                             # 'segment_b1': data.get('Segment B1')[-1],
#                             # 'segment_b2': data.get('Segment B2')[-1],
#                             # 'segment_c': data.get('Segment C')[-1],
#                             # 'format_info': data.get('Format Info')[-1],
#                             # 'version_info': data.get('Version Info')[-1],
#                             # 'contrast_uniformity': data.get('Contrast Uniformity')[-1],
#                             # 'x_growth': data.get('X Growth')[-1],
#                             # 'y_growth': data.get('Y Growth')[-1],
#                             # 'matrix_size': data.get('MatrixSize')[-1],
#                             # 'data': data.get('Data')[-1],
#                             # 'encoded_data': data.get('Encoded Data')[-1],
#                             #
#                         })
#                     print("BBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBB", data)
# # DELETE FILE  ##########################
# # ftp = ftplib.FTP(ip)
# # ftp.login(id, ps)
# # ftp.cwd(path)
# # ftp.delete(file)
# # ftp.close()
# #
# # path3 = '/home/maven/Music/'
# #
# # list_ = []
# # for file_ in path3:
# #     fileList = glob.glob(path3 + "/*.xlsx")
# #     fileList1 = " ".join(str(x) for x in fileList)
# #     try:
# #         os.remove(fileList1)
# #
# #     except Exception as e:
# #         print(e)
#
#
# # server = "192.168.1.4"
# # username = "ftp_maven"
# # password = "ftp_maven"
# # ftp = ftplib.FTP(server, username, password)
# # for root, dirs, files in os.walk('/home/maven/FTP FILE/excel'):
# #     for fname in files:
# #         print("22244", fname)
# #         a = fname[:2]
# #         print(")))))))", a)
# #         if a == "FTP":
# #         # print("22222222222",a)
# #             full_fname = os.path.join(root, fname)
# #             ftp.storbinary('STOR FTPUpload/' + fname, open(full_fname, 'rb'))
# #     else:
# #          pass
